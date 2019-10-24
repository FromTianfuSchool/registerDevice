'''
@File    :   uploadInfo.py
@start   :   2019/09/28 18:00:24
@Author  :   Jiang Xin
@Version :   1.0
@Contact :   gental_j@163.com
@License :   (C)Copyright 2018-2019, JiangXin
@Desc    :   upload used device's information
'''
import requests
from openpyxl import load_workbook, Workbook
import json
import datetime

url = "http://deviceman.caeri.com.cn:6037/Index/add_log.html"

# 扫码登陆随机数生成地址
url1 = "http://deviceman.caeri.com.cn:6037/Index/get_rand.html"
cookies = {
    "thinkphp_show_page_trace": "0|0",
    "PHPSESSID": "4bknp8svurumnlcoc1bao4t006"
}
headers = {
    "User-Agent":
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36",
}


def upload(dev_info="name"):
    """
    @param dev_info: data of device information
    """
    data_table = loadTemplate()
    load_data = generatePeriodData()
    for i in range(len(load_data)):
        for key in data_table.keys():
            data_table[key] = load_data[i][key]
        r = requests.post(url=url,
                          headers=headers,
                          cookies=cookies,
                          data=data_table)
        if r.status_code == 200 or r.status_code == '200':
            print('upload successful!')
    # print(data_table)


def loadDataExcel():
    """
     a list contains group data
     now replace by generatePeriodData()
    """
    wb = load_workbook('device_info.xlsx', data_only=True)
    workesheet = wb.active
    # rows = []
    rows_obj = list(workesheet.iter_rows())
    # print(rows_obj)
    # print(rows_obj[2][0])
    dict_list = []
    name_list = [cell.value for cell in rows_obj[2]]
    for row in rows_obj[3:]:
        if row[0].value:  # 空白行为数据录入结束标志位
            value_list = [cell.value for cell in row]
            group = dict(zip(name_list, value_list))
            group['data[start_time]'] = group[
                'data[start_time]'] + ' ' + group['data[start_hour]']
            dict_list.append(group)
    return dict_list
    # 发动机排气采样分析测试系统 ZPJ190 发动机排放检测部 2019-09-06 00:00:00 8 None 蒋鑫 None 13368385052 蒋鑫 车用压燃式发动机排气污染物 None

# loadDataExcel()


def generatePeriodData(period=1):
    '''
    read origin data from an excel, 
    if exist more than one row, return all, 
    if only one row exists, generate a period data by use input how long the devices used.  
    '''
    wb = load_workbook('device_info.xlsx', data_only=True)
    workesheet = wb.active
    allrows = list(workesheet.iter_rows())

    d_dict = []
    d_keys = [cell.value for cell in allrows[2]]

    if allrows[4][0].value:
        # if user input more than one line data, we use user' input
        for row in allrows[3:]:
            if row[0].value:
                # finish read excel when miss blank row
                d_values = [cell.value for cell in row]
                group = dict(zip(d_keys, d_values))
                group['data[start_time]'] = group[
                    'data[start_time]'] + ' ' + group['data[start_hour]']
                d_dict.append(group)
        return d_dict

    if allrows[3][0].value:
        d_values = [cell.value for cell in allrows[3]]
        group = dict(zip(d_keys, d_values))
        dates = getperiod(group['data[start_time]'])
        for date in dates:
            group['data[start_time]'] = date + ' ' + group['data[start_hour]']
            d_dict.append(group)
        return d_dict


def getperiod(startdate):
    '''
    use start date to genarate a period of date
    '''
    keyboard_in = input(
        'how many days you use those devices(it must be continuous!): ')
    if keyboard_in.isdigit():
        days = int(keyboard_in)

    dates = []

    for i in range(days+1):
        dt = datetime.datetime.strptime(startdate, '%Y-%m-%d')
        target_date = dt + datetime.timedelta(days=i)
        dates.append(target_date.strftime('%Y-%m-%d'))

    return dates


# startdate = '2019-09-06 00:00:00'
# getperiod(startdate=startdate)

# generatePeriodData()


def loadTemplate(template='format.json'):
    """
    description: get the parameter for upload data
    @param template: the path of template file 
    @return: dict
    """
    with open(template, 'r', encoding='utf-8') as rb:
        parm_dict = json.load(rb)

    return parm_dict


if __name__ == '__main__':

    # upload()
